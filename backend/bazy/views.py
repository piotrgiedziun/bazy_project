# -*- coding: utf-8 -*-
# django
from django.contrib.auth.models import User, Group
from django.shortcuts import get_list_or_404, render, redirect
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import Http404, HttpResponse
from django.conf import settings
from django.db.models import Sum, Avg, Max, Min
from django.utils.encoding import smart_str

# internal
from models import Newsy, Oplaty, Oplaty_type
from decorators import login_mieszkaniec_required

# python
from datetime import date
import itertools, os

# export
from openpyxl import Workbook

def home(request):
    return render(request, 'home.html', {'title': 'Strona główna'})

@login_mieszkaniec_required
def panel_rozliczenia(request):
    return render(request, 'panel/rozliczenia.html', {'title': 'Rozliczenia'})

@login_mieszkaniec_required
def panel_oplaty_chart_1(request):
    mieszkaniec = request.user.get_profile()
    oplaty_all = get_list_or_404(Oplaty, mieszkanie=mieszkaniec.mieszkanie)
    oplata_suma = Oplaty.objects.exclude(oplaty_type__name='nadpłata')\
        .filter(mieszkanie=mieszkaniec.mieszkanie).values('data_platnosci')\
        .annotate(sum=Sum('saldo'))

    oplaty_by_type = {}
    for o in oplaty_all:
        if not o.oplaty_type.pk in oplaty_by_type:
            oplaty_by_type[o.oplaty_type.pk] = []
        oplaty_by_type[o.oplaty_type.pk].append(o)

    return render(request, 'panel/oplaty_chart_1.html', {'title': 'Oplaty (wykres)', 'oplaty': oplaty_by_type.items(),'oplata_suma':oplata_suma})

@login_mieszkaniec_required
def panel_oplaty_chart_2(request):
    mieszkaniec = request.user.get_profile()

    oplaty = Oplaty.objects.exclude(oplaty_type__name='nadpłata')\
        .filter(mieszkanie=mieszkaniec.mieszkanie).values('oplaty_type__name')\
        .annotate(avg=Avg('saldo'), max=Max('saldo'), min=Min('saldo'), sum=Sum('saldo'))

    return render(request, 'panel/oplaty_chart_2.html', {'title': 'Oplaty (wykres)', 'oplaty': oplaty})

@login_mieszkaniec_required
def panel_oplaty(request):
    today = date.today()

    # date validation
    try:
        month = int(request.GET.get('month', today.month))
        year = int(request.GET.get('year', today.year))
    except ValueError:
        raise Http404("invalid parms")

    if not month in range(1, 13):
        raise Http404("invalid month")
    if not year in range(2000, today.year+1):
        raise Http404("invalid year")

    mieszkaniec = request.user.get_profile()

    # if does not exists -> 404
    oplaty = get_list_or_404(Oplaty,
        mieszkanie=mieszkaniec.mieszkanie,
        data_platnosci__month=month,
        data_platnosci__year=year,
    )

    pln_sum = sum([oplata.saldo for oplata in oplaty])

    # wyznaczanie nastpnych
    max_date = max([oplata.data_platnosci for oplata in oplaty])
    min_date = min([oplata.data_platnosci for oplata in oplaty])

    next = Oplaty.objects.filter(mieszkanie=mieszkaniec.mieszkanie, data_platnosci__gt=max_date)
    if len(next) >= 1:
        next = next[0]

    prev = Oplaty.objects.filter(mieszkanie=mieszkaniec.mieszkanie, data_platnosci__lt=min_date)
    if len(prev) >= 1:
        prev = prev[len(prev)-1]

    return render(request, 'panel/oplaty.html',
        {'title': 'Oplaty', 'oplaty': oplaty, 'pln_sum': pln_sum, 'next':next, 'prev':prev})

@login_mieszkaniec_required
def panel_komunikaty(request):
    mieszkaniec = request.user.get_profile()
    newsy = Newsy.objects.filter(mieszkancy=mieszkaniec)

    paginator = Paginator(newsy, settings.KOMUNIKATY_PER_PAGE)

    page = request.GET.get('page')
    try:
        newsy_paginate = paginator.page(page)
    except PageNotAnInteger:
        newsy_paginate = paginator.page(1)
    except EmptyPage:
        newsy_paginate = paginator.page(paginator.num_pages)

    return render(request, 'panel/komunikaty.html', {'title': 'Komunikaty', 'newsy': newsy_paginate, 'pages': paginator.page_range})

@login_mieszkaniec_required
def panel_komunikat(request, news_pk):
    mieszkaniec = request.user.get_profile()
    news = Newsy.objects.get(mieszkancy=mieszkaniec, pk=news_pk)

    # wyznaczanie nastpnych
    next = Newsy.objects.filter(mieszkancy=mieszkaniec, pk__gt=news_pk)
    if len(next) >= 1:
        next = next[len(next)-1]

    prev = Newsy.objects.filter(mieszkancy=mieszkaniec, pk__lt=news_pk)
    if len(prev) >= 1:
        prev = prev[0]

    return render(request, 'panel/komunikat.html', {'title': 'Komunikaty', 'news': news, 'next':next, 'prev':prev})

@login_mieszkaniec_required
def panel_export_main(request):
    mieszkaniec = request.user.get_profile()

    years = []
    for o in Oplaty.objects.filter(mieszkanie=mieszkaniec.mieszkanie).values('data_platnosci').annotate():
        if not o['data_platnosci'].year in years:
            years.append(o['data_platnosci'].year)

    return render(request, 'panel/export.html', {'title': 'Eksport', 'years':years})

@login_mieszkaniec_required
def panel_export(request, year):
    mieszkaniec = request.user.get_profile()

    oplaty_all = Oplaty.objects.filter(
        mieszkanie=mieszkaniec.mieszkanie,
        data_platnosci__year=year
    ).order_by('data_platnosci')

    oplaty_type = Oplaty_type.objects.all()

    grouped = itertools.groupby(oplaty_all, lambda record: record.data_platnosci.strftime("%m-%Y"))

    # generate
    wb = Workbook(encoding='utf-8')
    ws = wb.get_active_sheet()
    ws.title = u"Opłaty"

    # dict oplaty_type
    i = 0
    oplaty_type_to_id  = {}
    for ot in oplaty_type:
        oplaty_type_to_id[ot.pk] = i
        i += 1

    # set data
    oplaty_list = []
    dates = []
    for date, items in grouped:
        dates.append(date)
        append_list = [0]*len(oplaty_type)
        for o in list(items):
            try:
                append_list[oplaty_type_to_id[o.oplaty_type.pk]] = o
            except:
                pass
        oplaty_list.append(append_list)

    for y in range(0, len(oplaty_list)):
        oplaty = oplaty_list[y]
        for x in range(0, len(oplaty_type)):
            d = ws.cell(row = x+1, column = y+1)
            try:
                d.value = float(oplaty[x].saldo)
            except:
                d.value = 0

    # set titles
    i = 1
    for v in [o.name for o in oplaty_type]:
        ws.cell(row=i, column=0).value = v
        i += 1
    i = 1
    for v in dates:
        ws.cell(row=0, column=i).value = v
        i += 1

    # set size
    ws.column_dimensions['A'].width = 25

    wb.save('export.xlsx')

    response = HttpResponse(mimetype='application/xlsx')
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str('export.xlsx')
    response.write(open('export.xlsx', 'rb').read())

    return response

@login_mieszkaniec_required
def password_change_done(request):
    messages.success(request, 'Twoje nowe hasło zostało ustawione.')
    return redirect('panel')

@login_required
def logout(request):
    messages.info(request, 'Zostałeś wylogowany.')
    auth_logout(request)
    return redirect('home')